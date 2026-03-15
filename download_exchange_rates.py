#!/usr/bin/env python3
"""
download_exchange_rates.py
==========================
Descarga los tipos de cambio históricos mensuales (moneda local por USD)
al cierre de cada mes, desde enero 2000 hasta la fecha actual.

Países incluidos:
  Argentina, Chile, Uruguay, Colombia, Peru, Ecuador,
  Aruba, Curacao, Trinidad y Tobago, Venezuela,
  Mexico, Puerto Rico, Panama, Costa Rica,
  Martinique, French Guiana, Guadeloupe, USVI, Brazil

Fuentes de datos:
  1. Yahoo Finance (yfinance)  — monedas flotantes principales
  2. FRED (Federal Reserve)    — Brasil, México, Venezuela VEF (backup)
  3. Tasas fijas / hardcoded   — economías dolarizadas y paridades fijas

Instalación de dependencias:
  pip install yfinance pandas openpyxl pandas-datareader

Uso:
  python download_exchange_rates.py

Salida:
  tipos_de_cambio_historicos.xlsx   ← una hoja con todos los países
  tipos_de_cambio_historicos.csv    ← mismo contenido en CSV

Notas importantes:
  ─ Ecuador:        dolarizado desde ene-2000; tasa = 1.0
  ─ Panama:         Balboa (PAB) anclado 1:1 al USD; tasa = 1.0
  ─ Puerto Rico:    territorio EEUU; usa USD; tasa = 1.0
  ─ USVI:           territorio EEUU; usa USD; tasa = 1.0
  ─ Martinique,
    French Guiana,
    Guadeloupe:     territorios FR; usan EUR; columna = EUR/USD
  ─ Aruba (AWG):    anclado 1.79 AWG/USD desde 1986; mayormente fijo
  ─ Curacao (ANG):  anclado 1.79 ANG/USD; reemplazado por XCG en 2024
  ─ Venezuela:      datos pre-2018 = Bolivar Fuerte (VEF), oficial;
                    datos pos-ago-2018 = Bolivar Soberano (VES);
                    tasas paralelas/negras NO se incluyen
  ─ Argentina:      tasa oficial BNA; no incluye dólar blue/CCL/MEP
"""

import sys
import time
import logging
from datetime import date
from pathlib import Path

import pandas as pd

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Parámetros ────────────────────────────────────────────────────────────────
START_DATE  = "2000-01-01"
END_DATE    = date.today().strftime("%Y-%m-%d")
OUTPUT_STEM = "tipos_de_cambio_historicos"

# ── Seriales FRED (end-of-period monthly donde disponible) ───────────────────
# Descargados vía pandas_datareader; no requieren API key.
FRED_SERIES = {
    "DEXBZUS": "Brazil_BRL_FRED",        # BRL per USD (diario → resample M)
    "DEXMXUS": "Mexico_MXN_FRED",        # MXN per USD
    "DEXVZUS": "Venezuela_VEF_FRED",     # VEF per USD (hasta ~2015, oficial)
    "DEXCOLN": "Colombia_COP_FRED",      # COP per USD (alt name)
}

# ── Definición de países ──────────────────────────────────────────────────────
# Campos: (nombre_pais, codigo_moneda, yahoo_ticker, tipo, tasa_fija, fred_id)
#
# tipo = 'yahoo'  → descargar de Yahoo Finance (yfinance)
# tipo = 'usd'    → país usa USD oficialmente  (tasa = 1.0)
# tipo = 'fixed'  → paridad fija predefinida   (tasa = tasa_fija)
# tipo = 'eur'    → usa EUR; se descarga USDEUR=X (EUR por USD)
# tipo = 'multi'  → múltiples fuentes (Venezuela)
#
# fred_id: serie FRED como fuente alternativa (None si no aplica)

COUNTRIES = [
    # Nombre                 Moneda  Yahoo Ticker    Tipo     Fija   FRED_id
    ("Argentina",            "ARS",  "USDARS=X",     "yahoo", None,  None),
    ("Chile",                "CLP",  "USDCLP=X",     "yahoo", None,  None),
    ("Uruguay",              "UYU",  "USDUYU=X",     "yahoo", None,  None),
    ("Colombia",             "COP",  "USDCOP=X",     "yahoo", None,  "DEXCOLN"),
    ("Peru",                 "PEN",  "USDPEN=X",     "yahoo", None,  None),
    ("Ecuador",              "USD",  None,            "usd",   1.0,   None),
    ("Aruba",                "AWG",  "USDAWG=X",     "yahoo", None,  None),
    ("Curacao",              "ANG",  "USDANG=X",     "yahoo", None,  None),
    ("Trinidad y Tobago",    "TTD",  "USDTTD=X",     "yahoo", None,  None),
    ("Venezuela",            "VES",  "USDVES=X",     "multi", None,  "DEXVZUS"),
    ("Mexico",               "MXN",  "USDMXN=X",     "yahoo", None,  "DEXMXUS"),
    ("Puerto Rico",          "USD",  None,            "usd",   1.0,   None),
    ("Panama",               "PAB",  None,            "fixed", 1.0,   None),
    ("Costa Rica",           "CRC",  "USDCRC=X",     "yahoo", None,  None),
    ("Martinique",           "EUR",  "USDEUR=X",     "eur",   None,  None),
    ("French Guiana",        "EUR",  "USDEUR=X",     "eur",   None,  None),
    ("Guadeloupe",           "EUR",  "USDEUR=X",     "eur",   None,  None),
    ("USVI",                 "USD",  None,            "usd",   1.0,   None),
    ("Brazil",               "BRL",  "USDBRL=X",     "yahoo", None,  "DEXBZUS"),
]


# ── Descarga Yahoo Finance ────────────────────────────────────────────────────

def fetch_yahoo(ticker: str, start: str, end: str, retries: int = 3) -> pd.Series:
    """
    Descarga cotizaciones mensuales de Yahoo Finance.
    Retorna Series con índice PeriodIndex('M') y precios de cierre
    del último día hábil de cada mes.
    """
    try:
        import yfinance as yf
    except ImportError:
        log.error("yfinance no instalado.  Ejecute:  pip install yfinance")
        sys.exit(1)

    for attempt in range(1, retries + 1):
        try:
            tk = yf.Ticker(ticker)
            df = tk.history(
                start=start,
                end=end,
                interval="1mo",
                auto_adjust=False,
                actions=False,
            )
            if df.empty:
                log.warning(f"    {ticker}: sin datos en Yahoo Finance")
                return pd.Series(dtype=float, name=ticker)

            s = df["Close"].copy()
            s.index = pd.PeriodIndex(s.index.to_period("M"))
            s.name = ticker
            log.info(
                f"    Yahoo  {ticker}: {len(s):>4} meses  "
                f"({s.index[0]} → {s.index[-1]})"
            )
            return s

        except Exception as exc:
            log.warning(f"    {ticker} intento {attempt}/{retries}: {exc}")
            if attempt < retries:
                wait = 2 ** attempt
                log.info(f"    Esperando {wait}s …")
                time.sleep(wait)

    log.error(f"    {ticker}: descarga fallida tras {retries} intentos")
    return pd.Series(dtype=float, name=ticker)


# ── Descarga FRED ─────────────────────────────────────────────────────────────

def fetch_fred(series_id: str, start: str, end: str) -> pd.Series:
    """
    Descarga serie diaria de FRED y re-muestrea al último valor del mes.
    Requiere pandas_datareader:  pip install pandas-datareader
    """
    try:
        from pandas_datareader import data as pdr
    except ImportError:
        log.warning("pandas_datareader no instalado; se omite FRED.")
        return pd.Series(dtype=float, name=series_id)

    try:
        raw = pdr.DataReader(series_id, "fred", start=start, end=end)
        s = (
            raw.squeeze()
            .resample("ME")          # último día del mes
            .last()
            .dropna()
        )
        s.index = pd.PeriodIndex(s.index.to_period("M"))
        s.name = series_id
        log.info(
            f"    FRED   {series_id}: {len(s):>4} meses  "
            f"({s.index[0]} → {s.index[-1]})"
        )
        return s
    except Exception as exc:
        log.warning(f"    FRED {series_id}: {exc}")
        return pd.Series(dtype=float, name=series_id)


# ── Construcción del DataFrame ────────────────────────────────────────────────

def build_dataframe(start: str, end: str) -> pd.DataFrame:
    """
    Construye el DataFrame completo con todos los países.
    Índice: fecha de fin de mes (Timestamp).
    Columnas: 'Pais (Moneda/USD)'.
    """
    full_idx = pd.period_range(start=start, end=end, freq="M")
    df = pd.DataFrame(index=full_idx)
    df.index.name = "Mes"

    # Caché para evitar descargar el mismo ticker varias veces
    _yf_cache: dict[str, pd.Series] = {}
    _fr_cache: dict[str, pd.Series] = {}

    def get_yahoo(ticker):
        if ticker not in _yf_cache:
            log.info(f"  Descargando Yahoo Finance: {ticker}")
            _yf_cache[ticker] = fetch_yahoo(ticker, start, end)
        return _yf_cache[ticker]

    def get_fred(sid):
        if sid and sid not in _fr_cache:
            log.info(f"  Descargando FRED: {sid}")
            _fr_cache[sid] = fetch_fred(sid, start, end)
        return _fr_cache.get(sid, pd.Series(dtype=float))

    for country, currency, ticker, kind, fixed_rate, fred_id in COUNTRIES:
        col = f"{country} ({currency}/USD)"
        log.info(f"─ {country:<22} [{kind.upper()}]  moneda={currency}")

        if kind == "usd":
            df[col] = 1.0

        elif kind == "fixed":
            df[col] = fixed_rate

        elif kind in ("yahoo", "eur"):
            s = get_yahoo(ticker).reindex(full_idx)
            df[col] = s.values

        elif kind == "multi":
            # Venezuela: combina VES (Yahoo, post-2018) y VEF (FRED, pre-2018)
            s_yes = get_yahoo(ticker).reindex(full_idx)
            s_fred = get_fred(fred_id).reindex(full_idx) if fred_id else None
            combined = s_yes.copy()
            if s_fred is not None:
                mask_missing = combined.isna()
                combined[mask_missing] = s_fred[mask_missing]
            df[col] = combined.values
            log.info(
                f"    Venezuela combinado: {combined.notna().sum()} meses con datos"
            )

        # Fallback FRED si Yahoo no cubrió todo el rango
        if kind == "yahoo" and fred_id:
            missing = df[col].isna().sum()
            if missing > 0:
                log.info(
                    f"  {country}: {missing} meses sin dato en Yahoo; "
                    f"intentando FRED {fred_id} …"
                )
                s_fred = get_fred(fred_id).reindex(full_idx)
                mask = df[col].isna()
                df.loc[mask, col] = s_fred[mask].values
                filled = missing - df[col].isna().sum()
                if filled:
                    log.info(f"  {country}: FRED rellenó {filled} meses adicionales")

    # Convertir índice a Timestamps de fin de mes (más legible en Excel)
    df.index = df.index.to_timestamp("M")
    df.index.name = "Fecha (fin de mes)"
    return df


# ── Salida ────────────────────────────────────────────────────────────────────

def save_outputs(df: pd.DataFrame, stem: str) -> None:
    """Guarda el DataFrame en Excel y CSV."""
    csv_path  = Path(f"{stem}.csv")
    xlsx_path = Path(f"{stem}.xlsx")

    # CSV
    df.to_csv(csv_path, float_format="%.6f", date_format="%Y-%m-%d")
    log.info(f"Guardado: {csv_path.resolve()}")

    # Excel
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Tipo_de_Cambio", float_format="%.4f")

        ws = writer.sheets["Tipo_de_Cambio"]

        # Formato de fecha en columna A
        from openpyxl.styles import numbers as xl_num, Font, PatternFill, Alignment
        date_fmt = "YYYY-MM-DD"
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = date_fmt

        # Autoajuste de columnas
        for col_cells in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = max(
                max_len + 2, 14
            )

    log.info(f"Guardado: {xlsx_path.resolve()}")


def print_summary(df: pd.DataFrame) -> None:
    """Imprime resumen de cobertura."""
    divider = "─" * 72
    print(f"\n{divider}")
    print(f"{'País (moneda)':<38} {'Datos':>6} {'Desde':>10} {'Hasta':>10}")
    print(divider)
    for col in df.columns:
        valid = df[col].dropna()
        n = len(valid)
        if n > 0:
            desde = valid.index[0].strftime("%Y-%m")
            hasta = valid.index[-1].strftime("%Y-%m")
            pct = 100.0 * n / len(df)
            print(f"{col:<38} {n:>6,d}  {desde:>10} → {hasta:<10}  ({pct:.0f}%)")
        else:
            print(f"{col:<38} {'SIN DATOS':>8}")
    print(divider)
    print(
        f"\nTotal: {len(df):,d} meses  ×  {len(df.columns):,d} países  "
        f"({START_DATE} → {END_DATE})\n"
    )


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    log.info(f"Período: {START_DATE}  →  {END_DATE}")
    log.info(f"Países:  {len(COUNTRIES)}")

    df = build_dataframe(START_DATE, END_DATE)
    print_summary(df)
    save_outputs(df, OUTPUT_STEM)
    log.info("✓ Proceso completado.")


if __name__ == "__main__":
    main()
